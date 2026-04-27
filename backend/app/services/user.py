"""
用户与资源组业务逻辑服务。

Phase 4：移除 user_permission / user_resource_group 旧表引用。
权限现在通过 Role → role_permission 获取；资源组通过 UserGroup → group_resource_group 获取。
"""

from __future__ import annotations

import csv
import logging
from datetime import UTC, datetime
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import AppException, ConflictException, NotFoundException
from app.core.security import hash_password, validate_password_strength, verify_password
from app.models.instance import Instance
from app.models.role import UserGroup, group_resource_group, role_permission, user_group_member
from app.models.user import Permission, ResourceGroup, Users, instance_resource_group
from app.schemas.user import (
    ResourceGroupCreate,
    ResourceGroupUpdate,
    UserCreate,
    UserUpdate,
)

logger = logging.getLogger(__name__)

USER_IMPORT_HEADERS = [
    "username",
    "display_name",
    "email",
    "phone",
    "role",
    "manager_username",
    "manager_display_name",
    "employee_id",
    "department",
    "title",
    "user_groups",
    "is_active",
    "is_superuser",
    "password",
]

USER_IMPORT_HEADER_ALIASES = {
    "username": "username",
    "用户名": "username",
    "display_name": "display_name",
    "显示名称": "display_name",
    "姓名": "display_name",
    "email": "email",
    "邮箱": "email",
    "phone": "phone",
    "手机号": "phone",
    "电话": "phone",
    "role": "role",
    "角色": "role",
    "manager_username": "manager_username",
    "直属上级": "manager_username",
    "上级用户名": "manager_username",
    "manager_display_name": "manager_display_name",
    "直属上级显示名": "manager_display_name",
    "直属上级姓名": "manager_display_name",
    "上级显示名": "manager_display_name",
    "employee_id": "employee_id",
    "工号": "employee_id",
    "department": "department",
    "部门": "department",
    "title": "title",
    "职位": "title",
    "岗位": "title",
    "user_groups": "user_groups",
    "用户组": "user_groups",
    "is_active": "is_active",
    "启用": "is_active",
    "是否启用": "is_active",
    "is_superuser": "is_superuser",
    "超级管理员": "is_superuser",
    "是否超级管理员": "is_superuser",
    "password": "password",
    "密码": "password",
}

USER_IMPORT_TEMPLATE_EXAMPLE = {
    "username": "zhangsan",
    "display_name": "张三",
    "email": "zhangsan@example.com",
    "phone": "13800138000",
    "role": "developer",
    "manager_username": "admin",
    "manager_display_name": "超级管理员",
    "employee_id": "EMP1001",
    "department": "研发部",
    "title": "后端工程师",
    "user_groups": "开发组;默认组用户组",
    "is_active": "true",
    "is_superuser": "false",
    "password": "Sagitta@2026A",
}

USER_IMPORT_FIELD_DOCS = [
    {
        "field": "username",
        "required": "是",
        "description": "用户名，系统唯一键。导入时按它判断是新增还是更新。",
        "example": "zhangsan",
    },
    {
        "field": "display_name",
        "required": "否",
        "description": "显示名称/姓名。留空时新用户默认使用 username。",
        "example": "张三",
    },
    {
        "field": "email",
        "required": "否",
        "description": "邮箱地址。",
        "example": "zhangsan@example.com",
    },
    {
        "field": "phone",
        "required": "否",
        "description": "手机号。",
        "example": "13800138000",
    },
    {
        "field": "role",
        "required": "否",
        "description": "角色标识或角色中文名，必须已存在于系统中。",
        "example": "developer / 开发工程师",
    },
    {
        "field": "manager_username",
        "required": "否",
        "description": "直属上级的用户名。导入时优先按这个字段匹配系统中的用户。",
        "example": "admin",
    },
    {
        "field": "manager_display_name",
        "required": "否",
        "description": "直属上级的显示名称/姓名。当 manager_username 为空时，导入会按唯一显示名称匹配。",
        "example": "超级管理员",
    },
    {
        "field": "employee_id",
        "required": "否",
        "description": "工号。",
        "example": "EMP1001",
    },
    {
        "field": "department",
        "required": "否",
        "description": "部门名称。",
        "example": "研发部",
    },
    {
        "field": "title",
        "required": "否",
        "description": "职位/岗位。",
        "example": "后端工程师",
    },
    {
        "field": "user_groups",
        "required": "否",
        "description": "用户组标识或中文名，多个值可用分号/逗号分隔，且用户组必须已存在。",
        "example": "开发组;默认组用户组",
    },
    {
        "field": "is_active",
        "required": "否",
        "description": "是否启用，支持 true/false、1/0、是/否、启用/停用。",
        "example": "true",
    },
    {
        "field": "is_superuser",
        "required": "否",
        "description": "是否为超级管理员，支持 true/false、1/0、是/否。",
        "example": "false",
    },
    {
        "field": "password",
        "required": "否",
        "description": "仅创建新用户时使用；留空则使用导入弹窗填写的默认密码。密码至少 8 位，且必须同时包含数字、大写字母、小写字母和特殊字符。",
        "example": "Sagitta@2026A",
    },
]

def _normalize_bool(value: object, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "on", "是", "启用"}:
        return True
    if normalized in {"0", "false", "no", "n", "off", "否", "停用"}:
        return False
    return default


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _split_multi_value(raw: str) -> list[str]:
    normalized = (
        raw.replace("；", ";")
        .replace("，", ",")
        .replace("|", ",")
        .replace("\n", ",")
    )
    parts: list[str] = []
    for token in normalized.replace(";", ",").split(","):
        value = token.strip()
        if value:
            parts.append(value)
    return parts


class UserService:
    @staticmethod
    async def get_by_id(db: AsyncSession, user_id: int) -> Users | None:
        result = await db.execute(
            select(Users)
            .options(
                selectinload(Users.user_groups),
                selectinload(Users.role),
                selectinload(Users.manager),
            )
            .where(Users.id == user_id)
        )
        return result.scalar_one_or_none()

    @staticmethod
    async def get_by_username(db: AsyncSession, username: str) -> Users | None:
        result = await db.execute(
            select(Users)
            .options(
                selectinload(Users.role),
            )
            .where(Users.username == username)
        )
        return result.scalar_one_or_none()

    @staticmethod
    async def get_by_phone(db: AsyncSession, phone: str) -> Users | None:
        result = await db.execute(
            select(Users).where(Users.phone == phone, Users.is_active.is_(True))
        )
        return result.scalar_one_or_none()

    @staticmethod
    async def _apply_user_update(
        db: AsyncSession,
        user: Users,
        data: UserUpdate,
    ) -> None:
        if data.display_name is not None:
            user.display_name = data.display_name
        if data.email is not None:
            user.email = data.email
        if data.phone is not None:
            user.phone = data.phone
        if data.is_active is not None:
            user.is_active = data.is_active
        if data.is_superuser is not None:
            user.is_superuser = data.is_superuser
        if data.role_id is not None:
            user.role_id = data.role_id if data.role_id else None
        if data.manager_id is not None:
            user.manager_id = data.manager_id if data.manager_id else None
        if data.employee_id is not None:
            user.employee_id = data.employee_id
        if data.department is not None:
            user.department = data.department
        if data.title is not None:
            user.title = data.title
        if data.user_group_ids is not None:
            await db.execute(
                delete(user_group_member).where(user_group_member.c.user_id == user.id)
            )
            ugs_result = await db.execute(
                select(UserGroup.id).where(UserGroup.id.in_(data.user_group_ids))
            )
            for group_id in ugs_result.scalars().all():
                await db.execute(
                    user_group_member.insert().values(user_id=user.id, group_id=group_id)
                )

    @staticmethod
    async def list_users(
        db: AsyncSession,
        page: int = 1,
        page_size: int = 20,
        search: str | None = None,
        is_active: bool | None = None,
        role_ids: list[int] | None = None,
        user_group_ids: list[int] | None = None,
        departments: list[str] | None = None,
        titles: list[str] | None = None,
        statuses: list[bool] | None = None,
        user_ids: list[int] | None = None,
    ) -> tuple[int, list[Users]]:
        query = select(Users).options(
            selectinload(Users.user_groups),
            selectinload(Users.role),
            selectinload(Users.manager),
        )
        if user_group_ids:
            query = query.join(
                user_group_member,
                Users.id == user_group_member.c.user_id,
            ).where(user_group_member.c.group_id.in_(user_group_ids))
        if search:
            query = query.where(
                Users.username.ilike(f"%{search}%")
                | Users.display_name.ilike(f"%{search}%")
                | Users.email.ilike(f"%{search}%")
                | Users.phone.ilike(f"%{search}%")
            )
        if statuses:
            query = query.where(Users.is_active.in_(statuses))
        elif is_active is not None:
            query = query.where(Users.is_active.is_(is_active))
        if role_ids:
            query = query.where(Users.role_id.in_(role_ids))
        if departments:
            query = query.where(Users.department.in_(departments))
        if titles:
            query = query.where(Users.title.in_(titles))
        if user_ids:
            query = query.where(Users.id.in_(user_ids))

        query = query.distinct()
        total_result = await db.execute(select(func.count()).select_from(query.subquery()))
        total = total_result.scalar_one()

        query = query.offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(query)
        return total, list(result.scalars().all())

    @staticmethod
    async def create_user(db: AsyncSession, data: UserCreate, *, auto_commit: bool = True) -> Users:
        existing = await UserService.get_by_username(db, data.username)
        if existing:
            raise ConflictException(f"用户名 '{data.username}' 已存在")

        user = Users(
            username=data.username,
            password=hash_password(data.password),
            display_name=data.display_name or data.username,
            email=data.email,
            phone=data.phone,
            is_superuser=data.is_superuser,
            auth_type="local",
            role_id=data.role_id if data.role_id else None,
            manager_id=data.manager_id if data.manager_id else None,
            employee_id=data.employee_id,
            department=data.department,
            title=data.title,
            password_changed_at=datetime.now(UTC),
        )
        db.add(user)
        await db.flush()

        if data.user_group_ids:
            ug_ids_result = await db.execute(
                select(UserGroup.id).where(UserGroup.id.in_(data.user_group_ids))
            )
            for group_id in ug_ids_result.scalars().all():
                await db.execute(
                    user_group_member.insert().values(user_id=user.id, group_id=group_id)
                )

        if auto_commit:
            await db.commit()
        user = await UserService.get_by_id(db, user.id)
        logger.info("user_created: %s", data.username)
        return user

    @staticmethod
    async def update_user(
        db: AsyncSession,
        user_id: int,
        data: UserUpdate,
        *,
        auto_commit: bool = True,
    ) -> Users:
        user = await UserService.get_by_id(db, user_id)
        if not user:
            raise NotFoundException(f"用户 ID={user_id} 不存在")

        await UserService._apply_user_update(db, user, data)
        if auto_commit:
            await db.commit()
        user = await UserService.get_by_id(db, user.id)
        return user

    @staticmethod
    async def delete_user(db: AsyncSession, user_id: int) -> None:
        user = await UserService.get_by_id(db, user_id)
        if not user:
            raise NotFoundException(f"用户 ID={user_id} 不存在")
        await db.delete(user)
        await db.commit()

    @staticmethod
    async def change_password(
        db: AsyncSession,
        user_id: int,
        old_password: str,
        new_password: str,
    ) -> None:
        user = await UserService.get_by_id(db, user_id)
        if not user:
            raise NotFoundException("用户不存在")
        if not verify_password(old_password, user.password):
            from app.core.exceptions import AppException

            raise AppException("原密码错误", code=400)
        if verify_password(new_password, user.password):
            from app.core.exceptions import AppException

            raise AppException("新密码不能与当前密码相同", code=400)
        user.password = hash_password(new_password)
        user.password_changed_at = datetime.now(UTC)
        await db.commit()

    @staticmethod
    async def get_permissions(db: AsyncSession, user_id: int) -> list[str]:
        """v2: 获取用户权限（通过角色获取，不再查 user_permission）。"""
        user = await UserService.get_by_id(db, user_id)
        if not user or not user.role_id:
            return []
        result = await db.execute(
            select(Permission.codename)
            .join(role_permission, Permission.id == role_permission.c.permission_id)
            .where(role_permission.c.role_id == user.role_id)
        )
        return list(result.scalars().all())

    @staticmethod
    async def get_merged_permissions(
        db: AsyncSession, user_id: int, db_user: Users | None = None
    ) -> list[str]:
        """v2: 获取用户合并权限（角色权限，兼容无角色用户返回空列表）。"""
        if db_user is None:
            db_user = await UserService.get_by_id(db, user_id)
        if not db_user or not db_user.role_id:
            return []
        result = await db.execute(
            select(Permission.codename)
            .join(role_permission, Permission.id == role_permission.c.permission_id)
            .where(role_permission.c.role_id == db_user.role_id)
        )
        return sorted(result.scalars().all())

    @staticmethod
    async def grant_permissions(db: AsyncSession, user_id: int, perm_codes: list[str]) -> None:
        """v2: 授予权限（通过角色系统，将权限码添加到用户的角色）。

        如果用户没有角色，自动创建一个自定义角色并关联。
        """
        from sqlalchemy.dialects.postgresql import insert as pg_insert

        from app.models.role import Role

        user = await UserService.get_by_id(db, user_id)
        if not user:
            raise NotFoundException(f"用户 ID={user_id} 不存在")

        result = await db.execute(select(Permission).where(Permission.codename.in_(perm_codes)))
        perms = list(result.scalars().all())
        existing_codes = {p.codename for p in perms}

        for code in perm_codes:
            if code not in existing_codes:
                perm = Permission(codename=code, name=code)
                db.add(perm)
                perm_results = await db.execute(
                    select(Permission).where(Permission.codename == code)
                )
                new_perm = perm_results.scalar_one_or_none()
                if new_perm:
                    perms.append(new_perm)
        await db.flush()

        if not user.role_id:
            role = Role(
                name=f"user_{user_id}",
                name_cn=f"用户 {user_id} 专用角色",
                description=f"用户 {user.username} 的自动生成角色",
                is_system=False,
            )
            db.add(role)
            await db.flush()
            user.role_id = role.id

        role_id = user.role_id
        for perm in perms:
            stmt = (
                pg_insert(role_permission)
                .values(role_id=role_id, permission_id=perm.id)
                .on_conflict_do_nothing()
            )
            await db.execute(stmt)
        await db.commit()

    @staticmethod
    async def revoke_permissions(db: AsyncSession, user_id: int, perm_codes: list[str]) -> None:
        """v2: 撤销权限（从用户的角色中移除权限码）。"""
        result = await db.execute(select(Permission).where(Permission.codename.in_(perm_codes)))
        perm_ids = [p.id for p in result.scalars().all()]

        user = await UserService.get_by_id(db, user_id)
        if not user or not user.role_id or not perm_ids:
            return

        await db.execute(
            delete(role_permission).where(
                role_permission.c.role_id == user.role_id,
                role_permission.c.permission_id.in_(perm_ids),
            )
        )
        await db.commit()

    @staticmethod
    async def init_default_permissions(db: AsyncSession) -> None:
        """初始化标准权限定义。"""
        default_perms = [
            ("menu_dashboard", "Dashboard 菜单"),
            ("menu_sqlworkflow", "SQL 工单菜单"),
            ("menu_query", "在线查询菜单"),
            ("menu_schema", "数据字典菜单"),
            ("menu_ops", "运维工具菜单"),
            ("menu_monitor", "可观测中心菜单"),
            ("menu_system", "系统管理菜单"),
            ("menu_audit", "审计日志菜单"),
            ("sql_submit", "提交 SQL 工单"),
            ("sql_submit_high_risk", "提交高危 SQL 工单"),
            ("sql_review", "审核 SQL 工单"),
            ("sql_execute", "执行工单（自己）"),
            ("sql_execute_for_resource_group", "执行工单（资源组）"),
            ("query_submit", "提交查询"),
            ("query_applypriv", "申请查询权限"),
            ("query_review", "审核查询权限申请"),
            ("query_mgtpriv", "管理查询权限"),
            ("query_all_instances", "查询所有实例"),
            ("query_resource_group_instance", "查询资源组内实例"),
            ("process_view", "查看会话"),
            ("process_kill", "Kill 会话"),
            ("monitor_all_instances", "查看所有实例监控"),
            ("monitor_config_manage", "管理采集配置"),
            ("monitor_apply", "申请监控权限"),
            ("monitor_review", "审批监控权限"),
            ("monitor_alert_manage", "管理告警规则"),
            ("archive_apply", "申请数据归档"),
            ("archive_review", "审批数据归档"),
            ("archive_execute", "执行数据归档"),
            ("audit_user", "查看审计日志"),
            ("system_config_manage", "管理系统配置"),
            ("instance_manage", "管理实例"),
            ("resource_group_manage", "管理资源组"),
            ("user_manage", "管理用户"),
        ]
        for codename, name in default_perms:
            existing = await db.execute(select(Permission).where(Permission.codename == codename))
            if not existing.scalar_one_or_none():
                db.add(Permission(codename=codename, name=name))
        await db.commit()

    @staticmethod
    async def export_users(
        db: AsyncSession,
        search: str | None = None,
        is_active: bool | None = None,
        role_ids: list[int] | None = None,
        user_group_ids: list[int] | None = None,
        departments: list[str] | None = None,
        titles: list[str] | None = None,
        statuses: list[bool] | None = None,
        user_ids: list[int] | None = None,
    ) -> list[dict[str, str]]:
        _, users = await UserService.list_users(
            db,
            page=1,
            page_size=10000,
            search=search,
            is_active=is_active,
            role_ids=role_ids,
            user_group_ids=user_group_ids,
            departments=departments,
            titles=titles,
            statuses=statuses,
            user_ids=user_ids,
        )
        rows: list[dict[str, str]] = []
        for user in users:
            role_name = ""
            if user.role:
                role_name = user.role.name or user.role.name_cn
            rows.append(
                {
                    "username": user.username,
                    "display_name": user.display_name or "",
                    "email": user.email or "",
                    "phone": user.phone or "",
                    "role": role_name,
                    "manager_username": user.manager.username if user.manager else "",
                    "manager_display_name": user.manager.display_name if user.manager else "",
                    "employee_id": user.employee_id or "",
                    "department": user.department or "",
                    "title": user.title or "",
                    "user_groups": ";".join(
                        [(ug.name_cn or ug.name) for ug in (user.user_groups or [])]
                    ),
                    "is_active": "true" if user.is_active else "false",
                    "is_superuser": "true" if user.is_superuser else "false",
                    "password": "",
                }
            )
        return rows

    @staticmethod
    def build_user_export_file(
        rows: list[dict[str, str]],
        export_format: str = "xlsx",
    ) -> tuple[bytes, str, str]:
        fmt = export_format.lower()
        if fmt not in {"xlsx", "csv"}:
            raise AppException("导出格式仅支持 xlsx 或 csv", code=400)

        if fmt == "csv":
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(USER_IMPORT_HEADERS)
            for row in rows:
                writer.writerow([row.get(key, "") for key in USER_IMPORT_HEADERS])
            return (
                output.getvalue().encode("utf-8-sig"),
                "text/csv; charset=utf-8",
                "users_export.csv",
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(USER_IMPORT_HEADERS)
        for row in rows:
            ws.append([row.get(key, "") for key in USER_IMPORT_HEADERS])
        content = BytesIO()
        wb.save(content)
        return (
            content.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "users_export.xlsx",
        )

    @staticmethod
    def build_user_import_template(export_format: str = "xlsx") -> tuple[bytes, str, str]:
        fmt = export_format.lower()
        if fmt not in {"xlsx", "csv"}:
            raise AppException("模板格式仅支持 xlsx 或 csv", code=400)

        if fmt == "csv":
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(USER_IMPORT_HEADERS)
            writer.writerow([USER_IMPORT_TEMPLATE_EXAMPLE.get(col, "") for col in USER_IMPORT_HEADERS])
            return (
                output.getvalue().encode("utf-8-sig"),
                "text/csv; charset=utf-8",
                "users_import_template.csv",
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "UsersTemplate"
        ws.append(USER_IMPORT_HEADERS)
        ws.append([USER_IMPORT_TEMPLATE_EXAMPLE.get(col, "") for col in USER_IMPORT_HEADERS])
        note_ws = wb.create_sheet("字段说明")
        note_ws.append(["字段名", "是否必填", "填写说明", "示例值"])
        for item in USER_IMPORT_FIELD_DOCS:
            note_ws.append(
                [item["field"], item["required"], item["description"], item["example"]]
            )
        content = BytesIO()
        wb.save(content)
        return (
            content.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "users_import_template.xlsx",
        )

    @staticmethod
    async def import_users(
        db: AsyncSession,
        filename: str,
        content: bytes,
        default_password: str,
    ) -> dict[str, object]:
        suffix = Path(filename or "").suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            raise AppException("仅支持导入 .csv 或 .xlsx 文件", code=400)
        try:
            validate_password_strength(default_password)
        except ValueError as exc:
            raise AppException(str(exc), code=400) from exc

        import_headers, rows = UserService._parse_user_import_rows(content, suffix)
        if not rows:
            raise AppException("导入文件为空，请至少提供一条用户记录", code=400)

        created = 0
        updated = 0
        errors: list[dict[str, object]] = []
        for index, row in enumerate(rows, start=2):
            normalized_row = row["normalized"]
            raw_row = row["raw"]
            try:
                changed = await UserService._upsert_imported_user(db, normalized_row, default_password)
                if changed == "created":
                    created += 1
                else:
                    updated += 1
            except AppException as exc:
                errors.append(
                    {
                        "row": index,
                        "username": normalized_row.get("username", ""),
                        "error": exc.message,
                        "row_data": raw_row,
                    }
                )
            except Exception as exc:  # noqa: BLE001
                errors.append(
                    {
                        "row": index,
                        "username": normalized_row.get("username", ""),
                        "error": str(exc),
                        "row_data": raw_row,
                    }
                )

        await db.commit()
        return {
            "total": len(rows),
            "created": created,
            "updated": updated,
            "failed": len(errors),
            "import_headers": import_headers,
            "errors": errors,
        }

    @staticmethod
    def _parse_user_import_rows(
        content: bytes,
        suffix: str,
    ) -> tuple[list[str], list[dict[str, dict[str, str]]]]:
        if suffix == ".csv":
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            headers = reader.fieldnames or []
            if not headers:
                raise AppException("CSV 文件缺少表头", code=400)
            original_headers = [_normalize_cell(header) for header in headers]
            normalized_headers = [
                USER_IMPORT_HEADER_ALIASES.get(_normalize_cell(header), _normalize_cell(header))
                for header in headers
            ]
            rows: list[dict[str, dict[str, str]]] = []
            for raw_row in reader:
                row: dict[str, str] = {}
                raw_display_row: dict[str, str] = {}
                for original_key, normalized_key in zip(headers, normalized_headers, strict=False):
                    display_key = _normalize_cell(original_key)
                    cell_value = _normalize_cell(raw_row.get(original_key))
                    row[normalized_key] = cell_value
                    raw_display_row[display_key] = cell_value
                if any(row.values()):
                    rows.append({"normalized": row, "raw": raw_display_row})
            return original_headers, rows

        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active
        raw_headers = [_normalize_cell(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        if not raw_headers:
            raise AppException("Excel 文件缺少表头", code=400)
        headers = [USER_IMPORT_HEADER_ALIASES.get(header, header) for header in raw_headers]
        rows: list[dict[str, dict[str, str]]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            normalized_row: dict[str, str] = {}
            raw_display_row: dict[str, str] = {}
            for raw_header, normalized_header, value in zip(raw_headers, headers, values, strict=False):
                if not normalized_header:
                    continue
                cell_value = _normalize_cell(value)
                normalized_row[normalized_header] = cell_value
                raw_display_row[raw_header] = cell_value
            if any(normalized_row.values()):
                rows.append({"normalized": normalized_row, "raw": raw_display_row})
        return raw_headers, rows

    @staticmethod
    async def _upsert_imported_user(
        db: AsyncSession,
        row: dict[str, str],
        default_password: str,
    ) -> str:
        username = row.get("username", "").strip()
        if not username:
            raise AppException("username / 用户名 不能为空", code=400)

        role_id = await UserService._resolve_role_id(db, row.get("role", ""))
        manager_id = await UserService._resolve_manager_id(
            db,
            raw_manager_username=row.get("manager_username", ""),
            raw_manager_display_name=row.get("manager_display_name", ""),
        )
        user_group_ids = await UserService._resolve_user_group_ids(db, row.get("user_groups", ""))
        row_password = (row.get("password") or "").strip()

        existing = await UserService.get_by_username(db, username)
        if existing:
            await UserService._apply_user_update(
                db,
                existing,
                UserUpdate(
                    display_name=row.get("display_name"),
                    email=row.get("email"),
                    phone=row.get("phone"),
                    is_active=_normalize_bool(row.get("is_active"), default=existing.is_active),
                    is_superuser=_normalize_bool(
                        row.get("is_superuser"), default=existing.is_superuser
                    ),
                    role_id=role_id if row.get("role", "").strip() else 0,
                    manager_id=manager_id if (row.get("manager_username", "").strip() or row.get("manager_display_name", "").strip()) else 0,
                    employee_id=row.get("employee_id"),
                    department=row.get("department"),
                    title=row.get("title"),
                    user_group_ids=user_group_ids,
                ),
            )
            if row_password:
                existing.password = hash_password(row_password)
                # Mark import-reset passwords as "initial" so the user must change them on next login.
                existing.password_changed_at = existing.created_at
            return "updated"

        await UserService.create_user(
            db,
            UserCreate(
                username=username,
                password=row_password or default_password,
                display_name=row.get("display_name") or username,
                email=row.get("email", ""),
                phone=row.get("phone", ""),
                is_superuser=_normalize_bool(row.get("is_superuser"), default=False),
                role_id=role_id,
                manager_id=manager_id,
                employee_id=row.get("employee_id", ""),
                department=row.get("department", ""),
                title=row.get("title", ""),
                user_group_ids=user_group_ids,
            ),
            auto_commit=False,
        )
        return "created"

    @staticmethod
    async def _resolve_role_id(db: AsyncSession, raw_role: str) -> int | None:
        role_value = raw_role.strip()
        if not role_value:
            return None
        result = await db.execute(
            select(UserService._role_model().id).where(
                (UserService._role_model().name == role_value)
                | (UserService._role_model().name_cn == role_value)
            )
        )
        role_id = result.scalar_one_or_none()
        if role_id is None:
            raise AppException(f"角色 '{role_value}' 不存在", code=400)
        return role_id

    @staticmethod
    async def _resolve_manager_id(
        db: AsyncSession,
        raw_manager_username: str,
        raw_manager_display_name: str,
    ) -> int | None:
        manager_username = raw_manager_username.strip()
        manager_display_name = raw_manager_display_name.strip()
        if not manager_username and not manager_display_name:
            return None
        if manager_username:
            result = await db.execute(select(Users.id).where(Users.username == manager_username))
            manager_id = result.scalar_one_or_none()
            if manager_id is not None:
                return manager_id
            if not manager_display_name:
                raise AppException(f"直属上级用户名 '{manager_username}' 不存在", code=400)

        if manager_display_name:
            display_name_result = await db.execute(
                select(Users.id).where(Users.display_name == manager_display_name)
            )
            matched_ids = list(display_name_result.scalars().all())
            if len(matched_ids) == 1:
                return matched_ids[0]
            if len(matched_ids) > 1:
                raise AppException(
                    f"直属上级显示名 '{manager_display_name}' 匹配到多个用户，请填写 manager_username",
                    code=400,
                )

        raise AppException(
            "直属上级不存在，请填写有效的 manager_username 或唯一的 manager_display_name",
            code=400,
        )

    @staticmethod
    async def _resolve_user_group_ids(db: AsyncSession, raw_groups: str) -> list[int]:
        group_names = _split_multi_value(raw_groups)
        if not group_names:
            return []
        result = await db.execute(
            select(UserGroup.id, UserGroup.name, UserGroup.name_cn).where(
                (UserGroup.name.in_(group_names)) | (UserGroup.name_cn.in_(group_names))
            )
        )
        resolved = list(result.all())
        found_names = {
            value
            for row in resolved
            for value in (row.name, row.name_cn)
            if value
        }
        missing = [name for name in group_names if name not in found_names]
        if missing:
            raise AppException(f"用户组不存在: {', '.join(missing)}", code=400)
        return [row.id for row in resolved]

    @staticmethod
    def _role_model():
        from app.models.role import Role

        return Role


class ResourceGroupService:
    @staticmethod
    async def get_by_id(db: AsyncSession, rg_id: int) -> ResourceGroup | None:
        result = await db.execute(select(ResourceGroup).where(ResourceGroup.id == rg_id))
        return result.scalar_one_or_none()

    @staticmethod
    async def get_by_ids(db: AsyncSession, rg_ids: list[int]) -> list[ResourceGroup]:
        if not rg_ids:
            return []
        result = await db.execute(select(ResourceGroup).where(ResourceGroup.id.in_(rg_ids)))
        return list(result.scalars().all())

    @staticmethod
    async def list_groups(
        db: AsyncSession,
        page: int = 1,
        page_size: int = 20,
        search: str | None = None,
    ) -> tuple[int, list[ResourceGroup]]:
        query = select(ResourceGroup).options(
            selectinload(ResourceGroup.instances),
            selectinload(ResourceGroup.user_groups),
        )
        if search:
            query = query.where(
                ResourceGroup.group_name.ilike(f"%{search}%")
                | ResourceGroup.group_name_cn.ilike(f"%{search}%")
            )
        total_q = await db.execute(select(func.count()).select_from(query.subquery()))
        total = total_q.scalar_one()
        query = query.offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(query)
        return total, list(result.scalars().all())

    @staticmethod
    async def create(db: AsyncSession, data: ResourceGroupCreate) -> ResourceGroup:
        existing = await db.execute(
            select(ResourceGroup).where(ResourceGroup.group_name == data.group_name)
        )
        if existing.scalar_one_or_none():
            raise ConflictException(f"资源组 '{data.group_name}' 已存在")
        payload = data.model_dump(exclude={"instance_ids", "user_group_ids"})
        rg = ResourceGroup(**payload)
        db.add(rg)
        await db.flush()
        if data.instance_ids:
            valid_instance_ids = await db.execute(
                select(Instance.id).where(
                    Instance.id.in_(data.instance_ids),
                    Instance.is_active,
                )
            )
            for instance_id in valid_instance_ids.scalars().all():
                await db.execute(
                    instance_resource_group.insert().values(
                        instance_id=instance_id,
                        resource_group_id=rg.id,
                    )
                )
        if data.user_group_ids:
            valid_group_ids = await db.execute(
                select(UserGroup.id).where(UserGroup.id.in_(data.user_group_ids))
            )
            for group_id in valid_group_ids.scalars().all():
                await db.execute(
                    group_resource_group.insert().values(
                        group_id=group_id,
                        resource_group_id=rg.id,
                    )
                )
        await db.commit()
        await db.refresh(rg)
        return rg

    @staticmethod
    async def update(db: AsyncSession, rg_id: int, data: ResourceGroupUpdate) -> ResourceGroup:
        result = await db.execute(
            select(ResourceGroup)
            .options(
                selectinload(ResourceGroup.instances),
                selectinload(ResourceGroup.user_groups),
            )
            .where(ResourceGroup.id == rg_id)
        )
        rg = result.scalar_one_or_none()
        if not rg:
            raise NotFoundException(f"资源组 ID={rg_id} 不存在")
        payload = data.model_dump(exclude_none=True, exclude={"instance_ids", "user_group_ids"})
        for field, value in payload.items():
            setattr(rg, field, value)
        if data.instance_ids is not None:
            await db.execute(
                delete(instance_resource_group).where(
                    instance_resource_group.c.resource_group_id == rg_id
                )
            )
            valid_instance_ids = await db.execute(
                select(Instance.id).where(
                    Instance.id.in_(data.instance_ids),
                    Instance.is_active,
                )
            )
            for instance_id in valid_instance_ids.scalars().all():
                await db.execute(
                    instance_resource_group.insert().values(
                        instance_id=instance_id,
                        resource_group_id=rg_id,
                    )
                )
        if data.user_group_ids is not None:
            await db.execute(
                delete(group_resource_group).where(
                    group_resource_group.c.resource_group_id == rg_id
                )
            )
            valid_group_ids = await db.execute(
                select(UserGroup.id).where(UserGroup.id.in_(data.user_group_ids))
            )
            for group_id in valid_group_ids.scalars().all():
                await db.execute(
                    group_resource_group.insert().values(
                        group_id=group_id,
                        resource_group_id=rg_id,
                    )
                )
        await db.commit()
        await db.refresh(rg)
        return rg

    @staticmethod
    async def delete(db: AsyncSession, rg_id: int) -> None:
        rg = await ResourceGroupService.get_by_id(db, rg_id)
        if not rg:
            raise NotFoundException(f"资源组 ID={rg_id} 不存在")
        await db.delete(rg)
        await db.commit()

    @staticmethod
    async def get_member_count(db: AsyncSession, rg_id: int) -> int:
        """v2: 返回通过用户组关联的成员总数（去重）。"""
        from app.models.role import group_resource_group, user_group_member

        member_result = await db.execute(
            select(func.count(func.distinct(user_group_member.c.user_id)))
            .select_from(user_group_member)
            .join(
                group_resource_group,
                user_group_member.c.group_id == group_resource_group.c.group_id,
            )
            .where(group_resource_group.c.resource_group_id == rg_id)
        )
        return member_result.scalar_one()
