"""
角色与用户组业务逻辑服务。
"""

from __future__ import annotations

import csv
import logging
from io import BytesIO, StringIO
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import AppException, ConflictException, NotFoundException
from app.models.role import (
    Role,
    UserGroup,
    group_resource_group,
    role_permission,
    user_group_member,
)
from app.models.user import Permission, ResourceGroup, Users

if TYPE_CHECKING:
    pass

logger = logging.getLogger(__name__)

USER_GROUP_IMPORT_HEADERS = [
    "name",
    "name_cn",
    "description",
    "leader_username",
    "leader_display_name",
    "parent_name",
    "parent_name_cn",
    "members",
    "resource_groups",
    "is_active",
]

USER_GROUP_IMPORT_HEADER_ALIASES = {
    "name": "name",
    "组标识": "name",
    "用户组标识": "name",
    "name_cn": "name_cn",
    "中文名": "name_cn",
    "用户组中文名": "name_cn",
    "description": "description",
    "描述": "description",
    "说明": "description",
    "leader_username": "leader_username",
    "组长用户名": "leader_username",
    "leader_display_name": "leader_display_name",
    "组长显示名": "leader_display_name",
    "组长姓名": "leader_display_name",
    "parent_name": "parent_name",
    "父组标识": "parent_name",
    "上级组标识": "parent_name",
    "parent_name_cn": "parent_name_cn",
    "父组中文名": "parent_name_cn",
    "上级组中文名": "parent_name_cn",
    "members": "members",
    "组成员": "members",
    "成员": "members",
    "resource_groups": "resource_groups",
    "资源组": "resource_groups",
    "关联资源组": "resource_groups",
    "is_active": "is_active",
    "启用": "is_active",
    "是否启用": "is_active",
}

USER_GROUP_IMPORT_TEMPLATE_EXAMPLE = {
    "name": "dev_team",
    "name_cn": "开发组",
    "description": "负责核心业务研发",
    "leader_username": "zhangsan",
    "leader_display_name": "张三",
    "parent_name": "engineering_center",
    "parent_name_cn": "研发中心",
    "members": "zhangsan;lisi",
    "resource_groups": "mysql_prod;订单核心库",
    "is_active": "true",
}

USER_GROUP_IMPORT_FIELD_DOCS = [
    {
        "field": "name",
        "required": "是",
        "description": "用户组标识，系统唯一键。导入时按它判断是新增还是更新。",
        "example": "dev_team",
    },
    {
        "field": "name_cn",
        "required": "否",
        "description": "用户组中文名。",
        "example": "开发组",
    },
    {
        "field": "description",
        "required": "否",
        "description": "用户组描述。",
        "example": "负责核心业务研发",
    },
    {
        "field": "leader_username",
        "required": "否",
        "description": "组长用户名。导入时优先按这个字段匹配系统用户。",
        "example": "zhangsan",
    },
    {
        "field": "leader_display_name",
        "required": "否",
        "description": "组长显示名称/姓名。当 leader_username 为空时按唯一显示名匹配。",
        "example": "张三",
    },
    {
        "field": "parent_name",
        "required": "否",
        "description": "父组标识。导入时优先按它匹配已有用户组。",
        "example": "engineering_center",
    },
    {
        "field": "parent_name_cn",
        "required": "否",
        "description": "父组中文名。当 parent_name 为空时按唯一中文名匹配。",
        "example": "研发中心",
    },
    {
        "field": "members",
        "required": "否",
        "description": "组成员用户名或显示名，多个值可用分号/逗号分隔。更新时会整体覆盖该用户组成员。",
        "example": "zhangsan;lisi",
    },
    {
        "field": "resource_groups",
        "required": "否",
        "description": "资源组标识或中文名，多个值可用分号/逗号分隔，且资源组必须已存在并处于启用状态。",
        "example": "mysql_prod;订单核心库",
    },
    {
        "field": "is_active",
        "required": "否",
        "description": "是否启用，支持 true/false、1/0、是/否、启用/停用。",
        "example": "true",
    },
]


def _normalize_bool(value: object, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "on", "是", "启用"}:
        return True
    if normalized in {"0", "false", "no", "n", "off", "否", "停用"}:
        return False
    return default


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _split_multi_value(raw: str) -> list[str]:
    normalized = raw.replace("；", ";").replace("，", ",").replace("|", ",").replace("\n", ",")
    parts: list[str] = []
    for token in normalized.replace(";", ",").split(","):
        value = token.strip()
        if value:
            parts.append(value)
    return parts

BUILTIN_ROLES: list[dict] = [
    {
        "name": "superadmin",
        "name_cn": "超级管理员",
        "description": "拥有系统全部权限，is_superuser=True 时绕过一切检查",
        "permissions": [
            "menu_dashboard",
            "menu_sqlworkflow",
            "menu_query",
            "menu_schema",
            "menu_ops",
            "sql_submit",
            "sql_submit_high_risk",
            "sql_review",
            "sql_execute",
            "sql_execute_for_resource_group",
            "query_submit",
            "query_applypriv",
            "query_review",
            "query_mgtpriv",
            "query_all_instances",
            "query_resource_group_instance",
            "process_view",
            "process_kill",
            "menu_monitor",
            "monitor_all_instances",
            "monitor_config_manage",
            "monitor_apply",
            "monitor_review",
            "monitor_alert_manage",
            "archive_apply",
            "archive_review",
            "archive_execute",
            "audit_user",
            "menu_system",
            "menu_audit",
            "system_config_manage",
            "instance_manage",
            "resource_group_manage",
            "user_manage",
        ],
    },
    {
        "name": "dba",
        "name_cn": "全局DBA",
        "description": "全局 DBA，拥有 query_all_instances + monitor_all_instances",
        "permissions": [
            "menu_dashboard",
            "menu_sqlworkflow",
            "menu_query",
            "menu_schema",
            "menu_ops",
            "sql_submit",
            "sql_submit_high_risk",
            "sql_review",
            "sql_execute",
            "sql_execute_for_resource_group",
            "query_submit",
            "query_applypriv",
            "query_review",
            "query_mgtpriv",
            "query_all_instances",
            "process_view",
            "process_kill",
            "menu_monitor",
            "monitor_all_instances",
            "monitor_config_manage",
            "monitor_apply",
            "monitor_review",
            "monitor_alert_manage",
            "archive_apply",
            "archive_review",
            "archive_execute",
            "audit_user",
            "menu_audit",
            "instance_manage",
            "resource_group_manage",
        ],
    },
    {
        "name": "dba_group",
        "name_cn": "资源组DBA",
        "description": "资源组 DBA，运维权限但实例范围限于资源组（无 _all_instances 权限）",
        "permissions": [
            "menu_dashboard",
            "menu_sqlworkflow",
            "menu_query",
            "menu_schema",
            "menu_ops",
            "sql_submit",
            "sql_submit_high_risk",
            "sql_review",
            "sql_execute",
            "sql_execute_for_resource_group",
            "query_submit",
            "query_applypriv",
            "query_review",
            "query_mgtpriv",
            "query_resource_group_instance",
            "process_view",
            "process_kill",
            "menu_monitor",
            "monitor_config_manage",
            "monitor_apply",
            "monitor_review",
            "archive_apply",
            "archive_review",
            "archive_execute",
            "audit_user",
            "menu_audit",
            "instance_manage",
            "resource_group_manage",
        ],
    },
    {
        "name": "developer",
        "name_cn": "开发工程师",
        "description": "开发工程师，可提交工单和查询申请",
        "permissions": [
            "menu_dashboard",
            "menu_sqlworkflow",
            "menu_query",
            "menu_schema",
            "sql_submit",
            "query_submit",
            "query_applypriv",
            "archive_apply",
        ],
    },
]


class RoleService:
    @staticmethod
    async def get_by_id(db: AsyncSession, role_id: int) -> Role | None:
        result = await db.execute(
            select(Role).options(selectinload(Role.permissions)).where(Role.id == role_id)
        )
        return result.scalar_one_or_none()

    @staticmethod
    async def get_by_name(db: AsyncSession, name: str) -> Role | None:
        result = await db.execute(
            select(Role).options(selectinload(Role.permissions)).where(Role.name == name)
        )
        return result.scalar_one_or_none()

    @staticmethod
    async def list_roles(
        db: AsyncSession,
        page: int = 1,
        page_size: int = 50,
        is_active: bool | None = None,
    ) -> tuple[int, list[Role]]:
        query = select(Role).options(selectinload(Role.permissions))
        if is_active is not None:
            query = query.where(Role.is_active.is_(is_active))
        total_result = await db.execute(select(func.count()).select_from(query.subquery()))
        total = total_result.scalar_one()
        query = query.offset((page - 1) * page_size).limit(page_size).order_by(Role.id)
        result = await db.execute(query)
        return total, list(result.scalars().all())

    @staticmethod
    async def create_role(
        db: AsyncSession,
        name: str,
        name_cn: str = "",
        description: str = "",
        permission_codes: list[str] | None = None,
    ) -> Role:
        existing = await RoleService.get_by_name(db, name)
        if existing:
            raise ConflictException(f"角色 '{name}' 已存在")
        role = Role(name=name, name_cn=name_cn, description=description)
        db.add(role)
        await db.flush()

        if permission_codes:
            await RoleService._set_permissions(db, role, permission_codes)

        await db.commit()
        await db.refresh(role)
        logger.info("role_created: %s", name)
        return role

    @staticmethod
    async def update_role(
        db: AsyncSession,
        role_id: int,
        name_cn: str | None = None,
        description: str | None = None,
        is_active: bool | None = None,
        permission_codes: list[str] | None = None,
    ) -> Role:
        role = await RoleService.get_by_id(db, role_id)
        if not role:
            raise NotFoundException(f"角色 ID={role_id} 不存在")
        if name_cn is not None:
            role.name_cn = name_cn
        if description is not None:
            role.description = description
        if is_active is not None:
            role.is_active = is_active
        if permission_codes is not None:
            await RoleService._set_permissions(db, role, permission_codes)
        await db.commit()
        return await RoleService.get_by_id(db, role.id)

    @staticmethod
    async def delete_role(db: AsyncSession, role_id: int) -> None:
        role = await RoleService.get_by_id(db, role_id)
        if not role:
            raise NotFoundException(f"角色 ID={role_id} 不存在")
        if role.is_system:
            raise ConflictException(f"内置角色 '{role.name}' 不可删除")
        from sqlalchemy import update

        await db.execute(update(Users).where(Users.role_id == role_id).values(role_id=None))
        await db.delete(role)
        await db.commit()

    @staticmethod
    async def _set_permissions(db: AsyncSession, role: Role, codes: list[str]) -> None:
        await db.execute(delete(role_permission).where(role_permission.c.role_id == role.id))
        result = await db.execute(select(Permission).where(Permission.codename.in_(codes)))
        perms = list(result.scalars().all())
        existing_codes = {p.codename for p in perms}
        for code in codes:
            if code not in existing_codes:
                perm = Permission(codename=code, name=code)
                db.add(perm)
                perms.append(perm)
        await db.flush()
        for perm in perms:
            stmt = (
                pg_insert(role_permission)
                .values(role_id=role.id, permission_id=perm.id)
                .on_conflict_do_nothing()
            )
            await db.execute(stmt)

    @staticmethod
    async def init_builtin_roles(db: AsyncSession) -> None:
        """初始化 4 个内置角色及其权限码关联。先确保权限码存在。"""
        from app.services.user import UserService

        await UserService.init_default_permissions(db)
        for role_def in BUILTIN_ROLES:
            existing = await RoleService.get_by_name(db, role_def["name"])
            if not existing:
                role = Role(
                    name=role_def["name"],
                    name_cn=role_def["name_cn"],
                    description=role_def["description"],
                    is_system=True,
                )
                db.add(role)
                await db.flush()
                await RoleService._set_permissions(db, role, role_def["permissions"])
                await db.commit()
            else:
                await RoleService._set_permissions(db, existing, role_def["permissions"])
                await db.commit()


class UserGroupService:
    @staticmethod
    async def get_by_name(db: AsyncSession, name: str) -> UserGroup | None:
        result = await db.execute(select(UserGroup).where(UserGroup.name == name))
        return result.scalar_one_or_none()

    @staticmethod
    async def _get_active_resource_groups(
        db: AsyncSession, resource_group_ids: list[int]
    ) -> list[ResourceGroup]:
        if not resource_group_ids:
            return []

        result = await db.execute(
            select(ResourceGroup).where(
                ResourceGroup.id.in_(resource_group_ids),
                ResourceGroup.is_active.is_(True),
            )
        )
        groups = list(result.scalars().all())
        if len(groups) != len(set(resource_group_ids)):
            raise AppException("停用或不存在的资源组不能关联到用户组")
        return groups

    @staticmethod
    async def get_by_id(db: AsyncSession, group_id: int) -> UserGroup | None:
        result = await db.execute(
            select(UserGroup)
            .options(
                selectinload(UserGroup.leader),
                selectinload(UserGroup.parent),
                selectinload(UserGroup.members),
                selectinload(UserGroup.resource_groups),
            )
            .where(UserGroup.id == group_id)
        )
        return result.scalar_one_or_none()

    @staticmethod
    async def list_groups(
        db: AsyncSession,
        page: int = 1,
        page_size: int = 50,
        is_active: bool | None = None,
        parent_id: int | None = None,
        search: str | None = None,
        group_ids: list[int] | None = None,
        leader_ids: list[int] | None = None,
        parent_ids: list[int] | None = None,
        resource_group_ids: list[int] | None = None,
        statuses: list[bool] | None = None,
    ) -> tuple[int, list[UserGroup]]:
        query = select(UserGroup).options(
            selectinload(UserGroup.leader),
            selectinload(UserGroup.parent),
            selectinload(UserGroup.members),
            selectinload(UserGroup.resource_groups),
        )
        if is_active is not None:
            query = query.where(UserGroup.is_active.is_(is_active))
        if parent_id is not None:
            query = query.where(UserGroup.parent_id == parent_id)
        if search:
            query = query.where(
                UserGroup.name.ilike(f"%{search}%") | UserGroup.name_cn.ilike(f"%{search}%")
            )
        if group_ids:
            query = query.where(UserGroup.id.in_(group_ids))
        if leader_ids:
            query = query.where(UserGroup.leader_id.in_(leader_ids))
        if parent_ids:
            query = query.where(UserGroup.parent_id.in_(parent_ids))
        if statuses:
            query = query.where(UserGroup.is_active.in_(statuses))
        if resource_group_ids:
            query = (
                query.join(
                    group_resource_group,
                    UserGroup.id == group_resource_group.c.group_id,
                )
                .where(group_resource_group.c.resource_group_id.in_(resource_group_ids))
                .distinct()
            )
        total_result = await db.execute(select(func.count()).select_from(query.subquery()))
        total = total_result.scalar_one()
        query = query.offset((page - 1) * page_size).limit(page_size).order_by(UserGroup.id)
        result = await db.execute(query)
        return total, list(result.scalars().all())

    @staticmethod
    async def create_group(
        db: AsyncSession,
        name: str,
        name_cn: str = "",
        description: str = "",
        leader_id: int | None = None,
        parent_id: int | None = None,
        resource_group_ids: list[int] | None = None,
        member_ids: list[int] | None = None,
    ) -> UserGroup:
        existing = await db.execute(select(UserGroup).where(UserGroup.name == name))
        if existing.scalar_one_or_none():
            raise ConflictException(f"用户组 '{name}' 已存在")
        group = UserGroup(
            name=name,
            name_cn=name_cn,
            description=description,
            leader_id=leader_id,
            parent_id=parent_id,
        )
        db.add(group)
        await db.flush()

        if resource_group_ids:
            resource_groups = await UserGroupService._get_active_resource_groups(
                db, resource_group_ids
            )
            for rg in resource_groups:
                await db.execute(
                    group_resource_group.insert().values(group_id=group.id, resource_group_id=rg.id)
                )

        if member_ids:
            for uid in member_ids:
                await db.execute(user_group_member.insert().values(user_id=uid, group_id=group.id))

        await db.commit()
        await db.refresh(group)
        logger.info("user_group_created: %s", name)
        return group

    @staticmethod
    async def update_group(
        db: AsyncSession,
        group_id: int,
        name_cn: str | None = None,
        description: str | None = None,
        leader_id: int | None = None,
        parent_id: int | None = None,
        is_active: bool | None = None,
        resource_group_ids: list[int] | None = None,
        member_ids: list[int] | None = None,
    ) -> UserGroup:
        group = await UserGroupService.get_by_id(db, group_id)
        if not group:
            raise NotFoundException(f"用户组 ID={group_id} 不存在")
        if name_cn is not None:
            group.name_cn = name_cn
        if description is not None:
            group.description = description
        if leader_id is not None:
            group.leader_id = leader_id or None
        if parent_id is not None:
            group.parent_id = parent_id or None
        if is_active is not None:
            group.is_active = is_active

        if resource_group_ids is not None:
            await db.execute(
                delete(group_resource_group).where(group_resource_group.c.group_id == group_id)
            )
            resource_groups = await UserGroupService._get_active_resource_groups(
                db, resource_group_ids
            )
            for rg in resource_groups:
                await db.execute(
                    group_resource_group.insert().values(
                        group_id=group_id, resource_group_id=rg.id
                    )
                )

        if member_ids is not None:
            await db.execute(
                delete(user_group_member).where(user_group_member.c.group_id == group_id)
            )
            for uid in member_ids:
                await db.execute(user_group_member.insert().values(user_id=uid, group_id=group_id))

        await db.commit()
        await db.refresh(group)
        return group

    @staticmethod
    async def delete_group(db: AsyncSession, group_id: int) -> None:
        group = await UserGroupService.get_by_id(db, group_id)
        if not group:
            raise NotFoundException(f"用户组 ID={group_id} 不存在")
        await db.delete(group)
        await db.commit()

    @staticmethod
    async def get_members(db: AsyncSession, group_id: int) -> list[Users]:
        result = await db.execute(
            select(Users)
            .join(user_group_member, Users.id == user_group_member.c.user_id)
            .where(user_group_member.c.group_id == group_id)
        )
        return list(result.scalars().all())

    @staticmethod
    async def get_resource_groups(db: AsyncSession, group_id: int) -> list[ResourceGroup]:
        result = await db.execute(
            select(ResourceGroup)
            .join(
                group_resource_group, ResourceGroup.id == group_resource_group.c.resource_group_id
            )
            .where(group_resource_group.c.group_id == group_id)
        )
        return list(result.scalars().all())

    @staticmethod
    async def get_user_groups_for_resource_group(
        db: AsyncSession, resource_group_id: int
    ) -> list[UserGroup]:  # noqa: F821
        result = await db.execute(
            select(UserGroup)
            .join(group_resource_group, UserGroup.id == group_resource_group.c.group_id)
            .where(group_resource_group.c.resource_group_id == resource_group_id)
        )
        return list(result.scalars().all())

    @staticmethod
    async def export_groups(
        db: AsyncSession,
        search: str | None = None,
        is_active: bool | None = None,
        parent_id: int | None = None,
        group_ids: list[int] | None = None,
        leader_ids: list[int] | None = None,
        parent_ids: list[int] | None = None,
        resource_group_ids: list[int] | None = None,
        statuses: list[bool] | None = None,
    ) -> list[dict[str, str]]:
        _, groups = await UserGroupService.list_groups(
            db,
            page=1,
            page_size=10000,
            is_active=is_active,
            parent_id=parent_id,
            search=search,
            group_ids=group_ids,
            leader_ids=leader_ids,
            parent_ids=parent_ids,
            resource_group_ids=resource_group_ids,
            statuses=statuses,
        )
        rows: list[dict[str, str]] = []
        for group in groups:
            rows.append(
                {
                    "name": group.name,
                    "name_cn": group.name_cn or "",
                    "description": group.description or "",
                    "leader_username": group.leader.username if group.leader else "",
                    "leader_display_name": group.leader.display_name if group.leader else "",
                    "parent_name": group.parent.name if group.parent else "",
                    "parent_name_cn": group.parent.name_cn if group.parent else "",
                    "members": ";".join(
                        [(member.display_name or member.username) for member in (group.members or [])]
                    ),
                    "resource_groups": ";".join(
                        [
                            (resource_group.group_name_cn or resource_group.group_name)
                            for resource_group in (group.resource_groups or [])
                        ]
                    ),
                    "is_active": "true" if group.is_active else "false",
                }
            )
        return rows

    @staticmethod
    def build_group_export_file(
        rows: list[dict[str, str]],
        export_format: str = "xlsx",
    ) -> tuple[bytes, str, str]:
        fmt = export_format.lower()
        if fmt not in {"xlsx", "csv"}:
            raise AppException("导出格式仅支持 xlsx 或 csv", code=400)

        if fmt == "csv":
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(USER_GROUP_IMPORT_HEADERS)
            for row in rows:
                writer.writerow([row.get(key, "") for key in USER_GROUP_IMPORT_HEADERS])
            return (
                output.getvalue().encode("utf-8-sig"),
                "text/csv; charset=utf-8",
                "user_groups_export.csv",
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "UserGroups"
        ws.append(USER_GROUP_IMPORT_HEADERS)
        for row in rows:
            ws.append([row.get(key, "") for key in USER_GROUP_IMPORT_HEADERS])
        content = BytesIO()
        wb.save(content)
        return (
            content.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "user_groups_export.xlsx",
        )

    @staticmethod
    def build_group_import_template(export_format: str = "xlsx") -> tuple[bytes, str, str]:
        fmt = export_format.lower()
        if fmt not in {"xlsx", "csv"}:
            raise AppException("模板格式仅支持 xlsx 或 csv", code=400)

        if fmt == "csv":
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(USER_GROUP_IMPORT_HEADERS)
            writer.writerow(
                [USER_GROUP_IMPORT_TEMPLATE_EXAMPLE.get(col, "") for col in USER_GROUP_IMPORT_HEADERS]
            )
            return (
                output.getvalue().encode("utf-8-sig"),
                "text/csv; charset=utf-8",
                "user_groups_import_template.csv",
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "UserGroupsTemplate"
        ws.append(USER_GROUP_IMPORT_HEADERS)
        ws.append([USER_GROUP_IMPORT_TEMPLATE_EXAMPLE.get(col, "") for col in USER_GROUP_IMPORT_HEADERS])
        note_ws = wb.create_sheet("字段说明")
        note_ws.append(["字段名", "是否必填", "填写说明", "示例值"])
        for item in USER_GROUP_IMPORT_FIELD_DOCS:
            note_ws.append(
                [item["field"], item["required"], item["description"], item["example"]]
            )
        content = BytesIO()
        wb.save(content)
        return (
            content.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "user_groups_import_template.xlsx",
        )

    @staticmethod
    async def import_groups(
        db: AsyncSession,
        filename: str,
        content: bytes,
    ) -> dict[str, object]:
        suffix = Path(filename or "").suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            raise AppException("仅支持导入 .csv 或 .xlsx 文件", code=400)

        import_headers, rows = UserGroupService._parse_group_import_rows(content, suffix)
        if not rows:
            raise AppException("导入文件为空，请至少提供一条用户组记录", code=400)

        created = 0
        updated = 0
        errors: list[dict[str, object]] = []
        for index, row in enumerate(rows, start=2):
            normalized_row = row["normalized"]
            raw_row = row["raw"]
            try:
                changed = await UserGroupService._upsert_imported_group(db, normalized_row)
                if changed == "created":
                    created += 1
                else:
                    updated += 1
            except AppException as exc:
                errors.append(
                    {
                        "row": index,
                        "name": normalized_row.get("name", ""),
                        "error": exc.message,
                        "row_data": raw_row,
                    }
                )
            except Exception as exc:  # noqa: BLE001
                errors.append(
                    {
                        "row": index,
                        "name": normalized_row.get("name", ""),
                        "error": str(exc),
                        "row_data": raw_row,
                    }
                )

        await db.commit()
        return {
            "total": len(rows),
            "created": created,
            "updated": updated,
            "failed": len(errors),
            "import_headers": import_headers,
            "errors": errors,
        }

    @staticmethod
    def _parse_group_import_rows(
        content: bytes,
        suffix: str,
    ) -> tuple[list[str], list[dict[str, dict[str, str]]]]:
        if suffix == ".csv":
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            headers = reader.fieldnames or []
            if not headers:
                raise AppException("CSV 文件缺少表头", code=400)
            original_headers = [_normalize_cell(header) for header in headers]
            normalized_headers = [
                USER_GROUP_IMPORT_HEADER_ALIASES.get(_normalize_cell(header), _normalize_cell(header))
                for header in headers
            ]
            rows: list[dict[str, dict[str, str]]] = []
            for raw_row in reader:
                row: dict[str, str] = {}
                raw_display_row: dict[str, str] = {}
                for original_key, normalized_key in zip(headers, normalized_headers, strict=False):
                    display_key = _normalize_cell(original_key)
                    cell_value = _normalize_cell(raw_row.get(original_key))
                    row[normalized_key] = cell_value
                    raw_display_row[display_key] = cell_value
                if any(row.values()):
                    rows.append({"normalized": row, "raw": raw_display_row})
            return original_headers, rows

        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active
        raw_headers = [
            _normalize_cell(cell)
            for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        if not raw_headers:
            raise AppException("Excel 文件缺少表头", code=400)
        headers = [USER_GROUP_IMPORT_HEADER_ALIASES.get(header, header) for header in raw_headers]
        rows: list[dict[str, dict[str, str]]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            normalized_row: dict[str, str] = {}
            raw_display_row: dict[str, str] = {}
            for raw_header, normalized_header, value in zip(raw_headers, headers, values, strict=False):
                if not normalized_header:
                    continue
                cell_value = _normalize_cell(value)
                normalized_row[normalized_header] = cell_value
                raw_display_row[raw_header] = cell_value
            if any(normalized_row.values()):
                rows.append({"normalized": normalized_row, "raw": raw_display_row})
        return raw_headers, rows

    @staticmethod
    async def _upsert_imported_group(
        db: AsyncSession,
        row: dict[str, str],
    ) -> str:
        name = row.get("name", "").strip()
        if not name:
            raise AppException("name / 组标识 不能为空", code=400)

        leader_id = await UserGroupService._resolve_leader_id(
            db,
            raw_username=row.get("leader_username", ""),
            raw_display_name=row.get("leader_display_name", ""),
        )
        parent_id = await UserGroupService._resolve_parent_id(
            db,
            raw_name=row.get("parent_name", ""),
            raw_name_cn=row.get("parent_name_cn", ""),
            current_name=name,
        )
        member_ids = await UserGroupService._resolve_member_ids(db, row.get("members", ""))
        resource_group_ids = await UserGroupService._resolve_resource_group_ids(
            db, row.get("resource_groups", "")
        )

        existing = await UserGroupService.get_by_name(db, name)
        if existing:
            await UserGroupService.update_group(
                db,
                existing.id,
                name_cn=row.get("name_cn"),
                description=row.get("description"),
                leader_id=leader_id if (row.get("leader_username", "").strip() or row.get("leader_display_name", "").strip()) else 0,
                parent_id=parent_id if (row.get("parent_name", "").strip() or row.get("parent_name_cn", "").strip()) else 0,
                is_active=_normalize_bool(row.get("is_active"), default=existing.is_active),
                resource_group_ids=resource_group_ids,
                member_ids=member_ids,
            )
            return "updated"

        await UserGroupService.create_group(
            db,
            name=name,
            name_cn=row.get("name_cn", ""),
            description=row.get("description", ""),
            leader_id=leader_id,
            parent_id=parent_id,
            resource_group_ids=resource_group_ids,
            member_ids=member_ids,
        )
        if row.get("is_active", "").strip():
            await UserGroupService.update_group(
                db,
                (await UserGroupService.get_by_name(db, name)).id,  # type: ignore[union-attr]
                is_active=_normalize_bool(row.get("is_active"), default=True),
            )
        return "created"

    @staticmethod
    async def _resolve_leader_id(
        db: AsyncSession,
        raw_username: str,
        raw_display_name: str,
    ) -> int | None:
        username_value = raw_username.strip()
        if username_value:
            result = await db.execute(select(Users.id).where(Users.username == username_value))
            leader_id = result.scalar_one_or_none()
            if leader_id is None:
                raise AppException(f"组长用户不存在: {username_value}", code=400)
            return leader_id

        display_name_value = raw_display_name.strip()
        if not display_name_value:
            return None

        result = await db.execute(select(Users.id).where(Users.display_name == display_name_value))
        matched_ids = list(result.scalars().all())
        if not matched_ids:
            raise AppException(f"组长用户不存在: {display_name_value}", code=400)
        if len(matched_ids) > 1:
            raise AppException(f"组长显示名匹配到多个用户，请改用用户名: {display_name_value}", code=400)
        return matched_ids[0]

    @staticmethod
    async def _resolve_parent_id(
        db: AsyncSession,
        raw_name: str,
        raw_name_cn: str,
        current_name: str,
    ) -> int | None:
        name_value = raw_name.strip()
        if name_value:
            result = await db.execute(select(UserGroup.id).where(UserGroup.name == name_value))
            parent_id = result.scalar_one_or_none()
            if parent_id is None:
                raise AppException(f"父组不存在: {name_value}", code=400)
            if name_value == current_name:
                raise AppException("父组不能设置为自己", code=400)
            return parent_id

        name_cn_value = raw_name_cn.strip()
        if not name_cn_value:
            return None

        result = await db.execute(select(UserGroup.id, UserGroup.name).where(UserGroup.name_cn == name_cn_value))
        matched_rows = list(result.all())
        if not matched_rows:
            raise AppException(f"父组不存在: {name_cn_value}", code=400)
        if len(matched_rows) > 1:
            raise AppException(f"父组中文名匹配到多个用户组，请改用组标识: {name_cn_value}", code=400)
        parent_id, parent_name = matched_rows[0]
        if parent_name == current_name:
            raise AppException("父组不能设置为自己", code=400)
        return parent_id

    @staticmethod
    async def _resolve_member_ids(db: AsyncSession, raw_members: str) -> list[int]:
        member_values = _split_multi_value(raw_members)
        if not member_values:
            return []

        result = await db.execute(
            select(Users.id, Users.username, Users.display_name).where(
                (Users.username.in_(member_values)) | (Users.display_name.in_(member_values))
            )
        )
        rows = list(result.all())
        found_ids: list[int] = []
        found_values: set[str] = set()
        for user_id, username, display_name in rows:
            found_ids.append(user_id)
            found_values.add(username)
            if display_name:
                found_values.add(display_name)

        missing = [value for value in member_values if value not in found_values]
        if missing:
            raise AppException(f"组成员不存在: {', '.join(missing)}", code=400)
        return list(dict.fromkeys(found_ids))

    @staticmethod
    async def _resolve_resource_group_ids(db: AsyncSession, raw_resource_groups: str) -> list[int]:
        resource_group_values = _split_multi_value(raw_resource_groups)
        if not resource_group_values:
            return []

        result = await db.execute(
            select(ResourceGroup.id, ResourceGroup.group_name, ResourceGroup.group_name_cn).where(
                ResourceGroup.is_active.is_(True),
                (ResourceGroup.group_name.in_(resource_group_values))
                | (ResourceGroup.group_name_cn.in_(resource_group_values)),
            )
        )
        rows = list(result.all())
        found_ids: list[int] = []
        found_values: set[str] = set()
        for resource_group_id, group_name, group_name_cn in rows:
            found_ids.append(resource_group_id)
            found_values.add(group_name)
            if group_name_cn:
                found_values.add(group_name_cn)

        missing = [value for value in resource_group_values if value not in found_values]
        if missing:
            raise AppException(f"资源组不存在或未启用: {', '.join(missing)}", code=400)
        return list(dict.fromkeys(found_ids))
