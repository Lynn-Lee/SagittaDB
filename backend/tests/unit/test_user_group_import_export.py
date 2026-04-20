"""
用户组导入导出单元测试。
"""

from io import BytesIO
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

from app.core.exceptions import AppException
from app.services.role import UserGroupService


class TestUserGroupImportTemplate:
    def test_build_xlsx_template_contains_docs_sheet(self):
        content, media_type, filename = UserGroupService.build_group_import_template("xlsx")

        workbook = load_workbook(filename=BytesIO(content))

        assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert filename == "user_groups_import_template.xlsx"
        assert workbook.sheetnames == ["UserGroupsTemplate", "字段说明"]
        assert workbook["UserGroupsTemplate"]["A1"].value == "name"
        assert workbook["字段说明"]["A2"].value == "name"


class TestUserGroupImportParsing:
    def test_parse_csv_rows_supports_cn_headers(self):
        content = (
            "组标识,中文名,组长用户名,组成员,资源组,是否启用\n"
            "dev_team,开发组,zhangsan,zhangsan;lisi,mysql_prod;订单核心库,true\n"
        ).encode("utf-8-sig")

        headers, rows = UserGroupService._parse_group_import_rows(content, ".csv")

        assert headers == ["组标识", "中文名", "组长用户名", "组成员", "资源组", "是否启用"]
        assert rows[0]["normalized"] == {
            "name": "dev_team",
            "name_cn": "开发组",
            "leader_username": "zhangsan",
            "members": "zhangsan;lisi",
            "resource_groups": "mysql_prod;订单核心库",
            "is_active": "true",
        }


class TestUserGroupResolvers:
    @pytest.mark.asyncio
    async def test_resolve_parent_id_rejects_self_reference(self):
        db = AsyncMock()
        result = MagicMock()
        result.scalar_one_or_none.return_value = 12
        db.execute = AsyncMock(return_value=result)

        with pytest.raises(AppException) as exc_info:
            await UserGroupService._resolve_parent_id(
                db,
                raw_name="dev_team",
                raw_name_cn="",
                current_name="dev_team",
            )

        assert exc_info.value.message == "父组不能设置为自己"
